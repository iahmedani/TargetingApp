from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages
from .forms import UserRegisterForm, UserLoginForm
from django.contrib.auth.decorators import login_not_required
import io
import os
import time
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, JsonResponse
from .models import CPDataModel, TPMDataModel, TPMCSVData, CSVData, Sample, CPDataModel1, Sample1, TPM_SC_Data, TPM_EE_Data,FinalApproval # Adjust the models as needed

import pandas as pd
import numpy as np
from django.views import View
from django.http import HttpResponse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import transaction
import logging
from io import BytesIO
import pytz
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Case, When, F, IntegerField, Sum, Subquery, OuterRef, Q, ExpressionWrapper, FloatField, Value,  CharField
from django.db.models.functions import Cast, Greatest, Coalesce
import random
import json
import datetime 


TOTAL_CLUSTERS = int(os.getenv('TOTAL_CLUSTERS'))
MAX_SAMPLE_SIZE = int(os.getenv('MAX_SAMPLE_SIZE'))
MIN_ACCEPTABLE_SAMPLE_SIZE_PERCENT = float(os.getenv('MIN_ACCEPTABLE_SAMPLE_SIZE'))
MAX_SAMPLE_SIZE_PER_CLUSTER = int(os.getenv('MAX_SAMPLE_SIZE_PER_CLUSTER'))
INCLUSION_ERROR_THRESHOLD = float(os.getenv('INCLUSION_ERROR_THRESHOLD'))
EXCLUSION_ERROR_THRESHOLD = float(os.getenv('EXCLUSION_ERROR_THRESHOLD'))
MIN_EXCLUSION_ERROR_RECORDS = int(os.getenv('MIN_EXCLUSION_ERROR_RECORDS'))

@login_not_required
def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Account created for {username}!')
            return redirect('login')
    else:
        form = UserRegisterForm()
    return render(request, 'register.html', {'form': form})

@login_not_required
def user_login(request):
    if request.method == 'POST':
        form = UserLoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                return redirect('home')
            else:
                messages.error(request, 'Invalid username or password')
    else:
        form = UserLoginForm()
    return render(request, 'login.html', {'form': form})

def user_logout(request):
    logout(request)
    return redirect('login')

# Create your views here.
def home(request):
    return render(request, 'index.html')

def dqa(request):
    return render(request, 'temp.html', {'title': 'Data Quality Assessment'})

def import_data(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']
        import_type = request.POST.get('import_type')        
        # Generate the new file name with timestamp
        file_name, file_extension = os.path.splitext(csv_file.name)
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        new_file_name = f"{file_name}_{timestamp}{file_extension}"
        
        # Define the upload path
        upload_path = os.path.join('uploads', new_file_name)
        
        # Save the file to the 'uploads' folder
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'uploads'))
        filename = fs.save(new_file_name, csv_file)
        
        # Optionally, save the file information to your model
        # Assuming you have a field `file_path` in your model to store the path
        # model_instance = CPDataModel(file_path=upload_path)
        # model_instance.save()

        return HttpResponse(f"File uploaded successfully: {filename}")
    
    return render(request, 'upload_csv.html')

def sampling(request):
    return render(request, 'sample_gen.html', {'title': 'Generate Sample'})

def sampling_borderline(request):
    return render(request, 'sample_gen_borderline.html', {'title': 'Generate Borderline Sample'})

def error_check(request):
    return render(request, 'temp.html', {'title': 'Error Checking'})

def final_list(request):
    return render(request, 'final_list_gen1.html', {'title': 'Final List Generation'})

def reports(request):
    return render(request, 'temp.html', {'title': 'Reports'})

def upload_media(request):
    return render(request, 'moda_actions.html', {'title': 'MoDa Actions'})

def upload_sample_to_moda(request):
    return render(request, 'temp.html', {'title': 'Upload Sample to MODA'})

def moda_user_mgt(request):
    return render(request, 'temp.html', {'title': 'User Management'})



logger = logging.getLogger(__name__)

class CSVImportView(View):
    template_name = 'upload_csv.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        # Check if this is supposed to be a PUT request
        if request.POST.get('_method') == 'PUT':
            return self._handle_file_upload(request, is_update=True)
        else:
            return self._handle_file_upload(request, is_update=False)

    def _handle_file_upload(self, request, is_update):
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Please upload a CSV file.')
            return render(request, self.template_name)

        try:
            # Read CSV file
            df = pd.read_csv(csv_file)
            
            # Replace '/' with '_' in column names
            df.columns = df.columns.str.replace('-', '_')
            
            if 'ben_id' in df.columns:
                messages.error(request, 'Seems like wrong file is sent for process.')
                return render(request, self.template_name)
            
            # Process the DataFrame
            df, error_rows = self.process_dataframe(df, is_update)

            # Generate response CSV
            output = BytesIO()
            df.to_csv(output, index=False)
            output.seek(0)

            # Create HTTP response
            response = HttpResponse(output, content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="import_results.csv"'
            
            if error_rows:
                messages.warning(request, f'{len(error_rows)} rows had errors. Check the output CSV for details.')
            else:
                messages.success(request, 'All rows were processed successfully.')

            return response

        except Exception as e:
            logger.error(f"Error processing CSV: {str(e)}")
            messages.error(request, f'An error occurred while processing the CSV: {str(e)}')
            return render(request, self.template_name)

    def process_dataframe(self, df, is_update):
        # Add import_status column
        df['import_status'] = ''

        # # Convert 'Yes'/'No' to boolean
        # bool_columns = [
        #     'is_principal', 'cfac_consulted', 'ag_work', 'child_5', 'plw', 'vul'
        # ] + [f'cfac_Q{i}' for i in range(1, 14)] + [f'A{i}' for i in range(1, 14)] + \
        #   ['cfac_exclusion', 'exclusion_1']
        
        # for col in bool_columns:
        #     if col in df.columns:
        #         df[col] = df[col].map({'Yes': True, 'No': False})
        
        
        df.rename(columns={'KEY': 'key'}, inplace=True)
        
        df['observation'] = df['observation'].map({1:True, 2:False})
        df['ag_work'] = df['ag_work'].map({1:True, 2:False})
        # Convert date columns
        date_columns = ['data_assess', 'SubmissionDate','start','end','today','date_return']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce', utc=True)
        # Sum columns A1 to A13 and store in a new column
        df['CP_Calculation'] = df[['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 
                                'A8', 'A9', 'A10', 'A11', 'A12', 'A13']].sum(axis=1)

        # Convert integer columns
        int_columns = ['ben_age', 'p_age', 'alter_age', 'child_5Num', 'c1age', 'c2age', 'c3age', 'c4age', 'c5age',
                       'pbw_num', 'CFAC_Calculation', 'CP_Calculation', 'difference']
        for col in int_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')

        error_rows = []

        # Get valid field names from the model
        valid_fields = set(f.name for f in CPDataModel1._meta.get_fields())

        # Get the current timezone
        current_tz = timezone.get_current_timezone()

        # Process rows
        for index, row in df.iterrows():
            try:
                with transaction.atomic():
                    # Filter out any columns that don't exist in the model and handle NaN values
                    valid_data = {}
                    for k, v in row.to_dict().items():
                        if k in valid_fields and pd.notna(v):
                            if k in date_columns and isinstance(v, pd.Timestamp):
                                # Convert to UTC, then to the current timezone
                                v = v.tz_convert(current_tz)
                            valid_data[k] = v

                    if is_update:
                        try:
                            obj = CPDataModel1.objects.get(key=valid_data['key'])
                            for key, value in valid_data.items():
                                setattr(obj, key, value)
                            obj.save()
                            df.at[index, 'import_status'] = 'Updated'
                        except CPDataModel1.DoesNotExist:
                            df.at[index, 'import_status'] = 'Not found - Update skipped'
                    else:
                        obj, created = CPDataModel1.objects.get_or_create(
                            key=valid_data['key'],
                            defaults=valid_data
                        )
                        df.at[index, 'import_status'] = 'Created' if created else 'Already exists'
            except Exception as e:
                logger.error(f"Error processing row {index}: {str(e)}")
                df.at[index, 'import_status'] = f'Error: {str(e)}'
                error_rows.append(index)

        return df, error_rows

# URL configuration remains the same:
# path('import-csv/', CSVImportView.as_view(), name='csv_import'),

class TPMCSVImportView(View):
    template_name = 'upload_csv.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        # Check if this is supposed to be a PUT request
        if request.POST.get('_method') == 'PUT':
            return self._handle_file_upload(request, is_update=True)
        else:
            return self._handle_file_upload(request, is_update=False)

    def _handle_file_upload(self, request, is_update):
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Please upload a CSV file.')
            return render(request, self.template_name)

        try:
            # Read CSV file
            df = pd.read_csv(csv_file)
            
            # Replace '/' with '_' in column names
            df.columns = df.columns.str.replace('-', '_')
            
            if 'ben_id' not in df.columns:
                messages.error(request, 'Seems like wrong file is sent for process.')
                return render(request, self.template_name)
                
            # Process the DataFrame
            df, error_rows = self.process_dataframe(df, is_update)

            # Generate response CSV
            output = BytesIO()
            df.to_csv(output, index=False)
            output.seek(0)

            # Create HTTP response
            response = HttpResponse(output, content_type='text/csv', charset='utf-8-sig')
            response['Content-Disposition'] = 'attachment; filename="import_results.csv"'
            
            if error_rows:
                messages.warning(request, f'{len(error_rows)} rows had errors. Check the output CSV for details.')
            else:
                messages.success(request, 'All rows were processed successfully.')

            return response

        except Exception as e:
            logger.error(f"Error processing CSV: {str(e)}")
            messages.error(request, f'An error occurred while processing the CSV: {str(e)}')
            return render(request, self.template_name)

    def process_dataframe(self, df, is_update):
        # Add import_status column
        df['import_status'] = ''

        # Convert date columns
        date_columns = ['data_assess',  'SubmissionDate','start',	'end',	'today']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce', utc=True)
                
        df.rename(columns={'KEY': 'key'}, inplace=True)
        df.rename(columns={'SB_province': 'SB_province1', 'SB_district': 'SB_district1','SB_B_1':'SB_province', 'SB_B_2':'SB_district'}, inplace=True)
        df.rename(columns={'SB_province1': 'SB_B_1', 'SB_district':'SB_B_2'}, inplace=True)

        # Convert integer columns
        int_columns = ['TPM_Calculation']
        for col in int_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int16')

        error_rows = []

        # Get valid field names from the model
        valid_fields = set(f.name for f in TPM_SC_Data._meta.get_fields())

        # Get the current timezone
        current_tz = timezone.get_current_timezone()

        # Process rows
        for index, row in df.iterrows():
            try:
                with transaction.atomic():
                    # Filter out any columns that don't exist in the model and handle NaN values
                    valid_data = {}
                    for k, v in row.to_dict().items():
                        if k in valid_fields and pd.notna(v):
                            if k in date_columns and isinstance(v, pd.Timestamp):
                                # Convert to UTC, then to the current timezone
                                v = v.tz_convert(current_tz)
                            valid_data[k] = v

                    if is_update:
                        try:
                            obj = TPM_SC_Data.objects.get(key=valid_data['key'])
                            for key, value in valid_data.items():
                                setattr(obj, key, value)
                            obj.save()
                            df.at[index, 'import_status'] = 'Updated'
                        except TPM_SC_Data.DoesNotExist:
                            df.at[index, 'import_status'] = 'Not found - Update skipped'
                    else:
                        ben_id = valid_data['ben_id'].split('_')[1]
                        _Sample = Sample1.objects.get(ben_id=ben_id)
                        obj, created = TPM_SC_Data.objects.get_or_create(
                            key=valid_data['key'],
                            sample=_Sample,
                            defaults=valid_data
                        )
                        df.at[index, 'import_status'] = 'Created' if created else 'Already exists'
            except Exception as e:
                logger.error(f"Error processing row {index}: {str(e)}")
                df.at[index, 'import_status'] = f'Error: {str(e)}'
                error_rows.append(index)

        return df, error_rows


class TPMEEVImportView(View):
    template_name = 'upload_csv.html'

    def get(self, request):
        return render(request, self.template_name)

    def post(self, request):
        # Check if this is supposed to be a PUT request
        if request.POST.get('_method') == 'PUT':
            return self._handle_file_upload(request, is_update=True)
        else:
            return self._handle_file_upload(request, is_update=False)

    def _handle_file_upload(self, request, is_update):
        csv_file = request.FILES.get('csv_file')
        if not csv_file:
            messages.error(request, 'Please upload a CSV file.')
            return render(request, self.template_name)

        try:
            # Read CSV file
            df = pd.read_csv(csv_file)
            
            # Replace '/' with '_' in column names
            df.columns = df.columns.str.replace('-', '_')
            
            if 'cfac_Q4' in df.columns:
                messages.error(request, 'Seems like wrong file is sent for process.')
                return render(request, self.template_name)
                
            # Process the DataFrame
            df, error_rows = self.process_dataframe(df, is_update)

            # Generate response CSV
            output = BytesIO()
            df.to_csv(output, index=False)
            output.seek(0)

            # Create HTTP response
            response = HttpResponse(output, content_type='text/csv', charset='utf-8-sig')
            response['Content-Disposition'] = 'attachment; filename="import_results.csv"'
            
            if error_rows:
                messages.warning(request, f'{len(error_rows)} rows had errors. Check the output CSV for details.')
            else:
                messages.success(request, 'All rows were processed successfully.')

            return response

        except Exception as e:
            logger.error(f"Error processing CSV: {str(e)}")
            messages.error(request, f'An error occurred while processing the CSV: {str(e)}')
            return render(request, self.template_name)

    def process_dataframe(self, df, is_update):
        # Add import_status column
        df['import_status'] = ''

        # Convert date columns
        date_columns = ['data_assess',  'SubmissionDate','start',	'end',	'today', 'date_return']
        for col in date_columns:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce', utc=True)
                
        df.rename(columns={'KEY': 'key'}, inplace=True)

        # Convert integer columns
        int_columns = ['TPM_Calculation']
        for col in int_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int16')

        error_rows = []

        # Get valid field names from the model
        valid_fields = set(f.name for f in TPM_EE_Data._meta.get_fields())

        # Get the current timezone
        current_tz = timezone.get_current_timezone()

        # Process rows
        for index, row in df.iterrows():
            try:
                with transaction.atomic():
                    # Filter out any columns that don't exist in the model and handle NaN values
                    valid_data = {}
                    for k, v in row.to_dict().items():
                        if k in valid_fields and pd.notna(v):
                            if k in date_columns and isinstance(v, pd.Timestamp):
                                # Convert to UTC, then to the current timezone
                                v = v.tz_convert(current_tz)
                            valid_data[k] = v

                    if is_update:
                        try:
                            obj = TPM_EE_Data.objects.get(key=valid_data['key'])
                            for key, value in valid_data.items():
                                setattr(obj, key, value)
                            obj.save()
                            df.at[index, 'import_status'] = 'Updated'
                        except TPM_EE_Data.DoesNotExist:
                            df.at[index, 'import_status'] = 'Not found - Update skipped'
                    else:
                        # _Sample = Sample1.objects.get(ben_id=valid_data['ben_id'])
                        obj, created = TPM_EE_Data.objects.get_or_create(
                            key=valid_data['key'],
                            # sample=_Sample,
                            defaults=valid_data
                        )
                        df.at[index, 'import_status'] = 'Created' if created else 'Already exists'
            except Exception as e:
                logger.error(f"Error processing row {index}: {str(e)}")
                df.at[index, 'import_status'] = f'Error: {str(e)}'
                error_rows.append(index)

        return df, error_rows


class CSVDataCountView(View):
    def get(self, request):
        # Perform the aggregation
        counts = CPDataModel1.objects.exclude(
            assessmentType='Replacement Assessment'
        ).values(
            'SB_ao', 'SB_province', 'SB_district', 'SB_area', 'SB_nahia'
        ).annotate(
            count=Count('id'),
            sample=Case(
                When(count__gte=1000, then=MAX_SAMPLE_SIZE),
                default=Greatest(
                    Cast(F('count') * 0.2, IntegerField()),
                    1
                ),
                output_field=IntegerField()
            )
        ).order_by('SB_ao', 'SB_province', 'SB_district', 'SB_area', 'SB_nahia')
        

        # Prepare the result
        result = {
            'counts': list(counts),
            'total': CPDataModel1.objects.count()
        }

        # Return JSON response
        return JsonResponse(result, safe=False)
    




def select_unique_cfacs(queryset, num_cfacs):
    """
    Selects unique CFACs randomly from the given queryset.
    
    Args:
    queryset (QuerySet): The base queryset to select CFACs from.
    num_cfacs (int): The number of unique CFACs to select. Defaults to 27.
    
    Returns:
    list: A list of dictionaries containing the selected CFAC names and their counts.
    """
    cfac_counts = list(queryset.values('SB_cfac_name')
                       .annotate(count=Count('id'))
                       .filter(count__gte=10)
                       .order_by('-count'))
    
    if len(cfac_counts) <= num_cfacs:
        return cfac_counts
    return cfac_counts[:num_cfacs]


def select_unique_villages(queryset, num_villages):
    """
    Selects unique Villages randomly from the given queryset.
    
    Args:
    queryset (QuerySet): The base queryset to select Villages from.
    num_villages (int): The number of unique Villages to select. Defaults to 27.
    
    Returns:
    list: A list of dictionaries containing the selected Villages names and their counts.
    """
    villages = list(queryset.values('SB_B_3')
                    .annotate(count=Count('id'))
                    .filter(count__gte=10)
                    .order_by('-count'))
    
    if len(villages) <= num_villages:
        return villages
    
    return villages[:num_villages]

# class GenerateSampleView(View):
#     def post(self, request, *args, **kwargs):
#         try:
#             data = json.loads(request.body)
#         except json.JSONDecodeError:
#             return JsonResponse({'error': 'Invalid JSON'}, status=400)

#         area_office = data.get('area_office')
#         province = data.get('province')
#         district = data.get('district')
#         nahia = data.get('nahia')
#         if nahia == 'null' or nahia is None:
#             nahia = None  # Assign None to nahia if 'null' is passed
#         sample_size = data.get('sampleSize')

#         if not all([area_office, province, district, sample_size]):
#             return JsonResponse({'error': 'Missing required parameters'}, status=400)

#         try:
#             sample_size = int(sample_size)
#         except ValueError:
#             return JsonResponse({'error': 'sampleSize must be an integer'}, status=400)
#         query = Q(SB_ao=area_office, SB_province=province, SB_district=district, vul="Yes") & ~Q(assessmentType='Replacement')
        
#         if nahia is not None:
#             query &= Q(SB_nahia=nahia)
#             sample_exists = Sample1.objects.filter(cp_id__in=CPDataModel1.objects.filter(query)).exists()
            
#             if sample_exists:
#                 print('sample already exists')
#                 return JsonResponse({'error': 'Sample already exists for the given parameters'}, status=400)
#             else:
#                 queryset = CPDataModel1.objects.filter(query)
#             # Urban area: Simple random sampling
#                 sample = list(queryset.filter(SB_nahia=nahia).order_by('?')[:sample_size])
#                 message = "Urban area: Simple random sampling based on nahia."
#                 instruction = "For exlusion Error, Please do one interview of the household after three (3) spot-check interviews, distribute 100 interviews to ensure maximum possible randomness. Total target for district is 100"
#         else:
            
#             sample_exists = Sample1.objects.filter(cp_id__in=CPDataModel1.objects.filter(query)).exists()
#             if sample_exists:
#                 return JsonResponse({'error': 'Sample already exists for the given parameters'}, status=400)
#             else:
#                 queryset = CPDataModel1.objects.filter(query)
#             # Rural area: Apply priority-based sampling
#             cfacs = select_unique_cfacs(queryset, num_cfacs=TOTAL_CLUSTERS)
            
#             if len(cfacs) == TOTAL_CLUSTERS:
#                 # Priority 1: CFAC-based sampling
#                 sample = set()
                
#                 for cfac in cfacs:
#                     cfac_records = set(queryset.filter(SB_cfac_name=cfac['SB_cfac_name']).order_by('?').values_list('id', flat=True)[:min(10, sample_size // TOTAL_CLUSTERS)])
#                     sample.update(cfac_records)
                
#                 message = f"Rural area: CFAC-based sampling with exactly {TOTAL_CLUSTERS} CFACs."
#                 instruction = "For exlusion Error, Please do 4 interviews per CFAC, you can use systematic sampling/random sampling to to ensure randomness. Total target for district is 100"
                
#             else:
#                 # Check for villages
#                 villages = select_unique_villages(queryset, num_villages=TOTAL_CLUSTERS)
                
#                 if len(villages) == TOTAL_CLUSTERS:
#                     # Priority 2: Village-based sampling
#                     sample = set()
                    
#                     for village in villages:
#                         village_records = set(queryset.filter(SB_B_3=village['SB_B_3']).order_by('?').values_list('id', flat=True)[:min(10, sample_size // TOTAL_CLUSTERS)])
#                         sample.update(village_records)
                    
#                     message = f"Rural area: Village-based sampling with exactly {TOTAL_CLUSTERS} villages."
#                     instruction = "For exlusion Error, Please do 4 interviews per Village, you can use systematic sampling/random sampling to to ensure randomness. Total target for district is 100"
#                 else:
#                     # Priority 3: Random sampling
#                     sample = set(queryset.order_by('?').values_list('id', flat=True)[:sample_size])
#                     message = f"Rural area: Random sampling due to not having exactly {TOTAL_CLUSTERS} CFACs or villages with 11+ records."
#                     instruction = "For exlusion Error, Please do one interview of the household after three (3) spot-check interviews, distribute 100 interviews to ensure maximum possible randomness. Total target for district is 100"
#             # Ensure we have the correct sample size
#             if len(sample) < sample_size:
#                 additional_records = set(queryset.exclude(id__in=sample).order_by('?').values_list('id', flat=True)[:sample_size - len(sample)])
#                 sample.update(additional_records)
            
#             # Convert set of IDs back to queryset
#             sample = queryset.filter(id__in=sample)

#         # Ensure the sample size does not exceed the requested size
#         sample = sample[:sample_size]

#         # Prepare the response data dynamically
#         sample_data = []
#         for record in sample:
#             record_data = {}
#             for field in record._meta.fields:
#                 field_name = field.name
#                 field_value = getattr(record, field_name)
#                 # Convert non-JSON serializable types to string
#                 if isinstance(field_value, (datetime.date, datetime.datetime)):
#                     field_value = field_value.isoformat()
#                 record_data[field_name] = field_value
#             sample_data.append(record_data)

#         response_data = {
#             "message": message,
#             "sample_size": len(sample_data),
#             "sample": sample_data,
#             "instruction": instruction
#         }

#         return JsonResponse(response_data, safe=False)    

from django.views import View
from django.http import JsonResponse
from django.db.models import Q
import json
import datetime

class GenerateSampleView(View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        area_office = data.get('area_office')
        province = data.get('province')
        district = data.get('district')
        nahia = data.get('nahia')
        sample_size = data.get('sampleSize')

        # Validate input parameters
        if nahia == 'null' or nahia is None:
            nahia = None
        
        if not all([area_office, province, district, sample_size]):
            return JsonResponse({'error': 'Missing required parameters'}, status=400)
        
        try:
            sample_size = int(sample_size)
        except ValueError:
            return JsonResponse({'error': 'sampleSize must be an integer'}, status=400)

        # Base query
        query = Q(SB_ao=area_office, SB_province=province, SB_district=district, vul="Yes") & ~Q(assessmentType='Replacement')
        

        # Add nahia to query if present
        if nahia is not None:
            query &= Q(SB_nahia=nahia)
            
        # Check if a sample already exists for this query
        sample_exists = Sample1.objects.filter(cp_id__in=CPDataModel1.objects.filter(query)).exists()
        
        if sample_exists:
            return JsonResponse({'error': 'Sample already exists for the given parameters'}, status=400)

        # Fetch the queryset based on the filter
        queryset = CPDataModel1.objects.filter(query)

        # Sampling logic
        if nahia is not None:
            # Urban area: Simple random sampling
            sample = list(queryset.order_by('?')[:sample_size])
            message = "Urban area: Simple random sampling based on nahia."
            instruction = "For exclusion error, please do one interview of the household after three (3) spot-check interviews. Distribute 100 interviews to ensure maximum randomness. Total target for district is 100."
        else:
            # Rural area: Priority-based sampling
            sample, message, instruction = self.sample_rural_area(queryset, sample_size)

        # Ensure the sample size does not exceed the requested size
        sample = sample[:sample_size]

        # Prepare the response data dynamically
        sample_data = self.prepare_sample_data(sample)

        # Response data
        response_data = {
            "message": message,
            "sample_size": len(sample_data),
            "sample": sample_data,
            "instruction": instruction
        }

        return JsonResponse(response_data, safe=False)

    def sample_rural_area(self, queryset, sample_size):
        TOTAL_CLUSTERS = 25  # Example value for total clusters
        cfacs = select_unique_cfacs(queryset, num_cfacs=TOTAL_CLUSTERS)

        if len(cfacs) == TOTAL_CLUSTERS:
            # CFAC-based sampling
            sample = set()
            for cfac in cfacs:
                cfac_records = set(queryset.filter(SB_cfac_name=cfac['SB_cfac_name']).order_by('?').values_list('id', flat=True)[:min(10, sample_size // TOTAL_CLUSTERS)])
                sample.update(cfac_records)

            message = f"Rural area: CFAC-based sampling with exactly {TOTAL_CLUSTERS} CFACs."
            instruction = "For exclusion error, please do 4 interviews per CFAC. You can use systematic/random sampling to ensure randomness. Total target for district is 100."
        else:
            villages = select_unique_villages(queryset, num_villages=TOTAL_CLUSTERS)

            if len(villages) == TOTAL_CLUSTERS:
                # Village-based sampling
                sample = set()
                for village in villages:
                    village_records = set(queryset.filter(SB_B_3=village['SB_B_3']).order_by('?').values_list('id', flat=True)[:min(10, sample_size // TOTAL_CLUSTERS)])
                    sample.update(village_records)

                message = f"Rural area: Village-based sampling with exactly {TOTAL_CLUSTERS} villages."
                instruction = "For exclusion error, please do 4 interviews per village. You can use systematic/random sampling. Total target for district is 100."
            else:
                # Random sampling
                sample = set(queryset.order_by('?').values_list('id', flat=True)[:sample_size])
                message = f"Rural area: Random sampling due to lack of exactly {TOTAL_CLUSTERS} CFACs or villages with 11+ records."
                instruction = "For exclusion error, please do one interview of the household after three (3) spot-check interviews. Distribute 100 interviews to ensure maximum randomness. Total target for district is 100."

        # Ensure we have the correct sample size
        if len(sample) < sample_size:
            additional_records = set(queryset.exclude(id__in=sample).order_by('?').values_list('id', flat=True)[:sample_size - len(sample)])
            sample.update(additional_records)

        # Convert set of IDs back to queryset
        sample = queryset.filter(id__in=sample)
        return sample, message, instruction

    def prepare_sample_data(self, sample):
        sample_data = []
        for record in sample:
            record_data = {}
            for field in record._meta.fields:
                field_name = field.name
                field_value = getattr(record, field_name)
                # Convert non-JSON serializable types to string
                if isinstance(field_value, (datetime.date, datetime.datetime)):
                    field_value = field_value.isoformat()
                record_data[field_name] = field_value
            sample_data.append(record_data)
        return sample_data

class ApproveSampleView(View):
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)

        sample_ids = data.get('ids')
        if not sample_ids:
            return JsonResponse({'error': 'Missing ids parameter'}, status=400)
        
        is_urban = data.get('is_urban')
        if is_urban is None:
            return JsonResponse({'error': 'Missing isUrban parameter'}, status=400)
        
        sample_type = data.get('sample_type')
        if not sample_type:
            return JsonResponse({'error': 'Missing sampleType parameter'}, status=400)
        
        remarks = data.get('remarks')
        if not remarks:
            return JsonResponse({'error': 'Missing remarks parameter'}, status=400)
        
        # Get CSVData objects for the given sample_ids
        csv_data_objects = CPDataModel1.objects.filter(id__in=sample_ids)
        
        if not csv_data_objects.exists():
            return JsonResponse({'error': 'No data in CPDataModel1 found for the given ids'}, status=404)
        
        # Prepare base query for existing samples
        existing_csv_data = CPDataModel1.objects.filter(
            SB_ao__in=csv_data_objects.values('SB_ao'),
            SB_province__in=csv_data_objects.values('SB_province'),
            SB_district__in=csv_data_objects.values('SB_district')
        )

        if is_urban & is_urban ==1:
            # For urban samples, include nahia in the query
            existing_csv_data = existing_csv_data.filter(
                SB_nahia__in=csv_data_objects.values('SB_nahia')
            )
            is_urban = True  # Set is_urban to True for urban samples
        else:
            is_urban = False  # Set is_urban to False for rural samples
            # For rural samples, exclude nahia from the query
        
        # Check if there are any existing samples
        if sample_type == 'Regular':
            existing_samples = Sample1.objects.filter(cp_id__in=existing_csv_data).exists()
        elif sample_type == 'Borderline':
            existing_samples = Sample1.objects.filter(cp_id__in=existing_csv_data, sample_type='Borderline').exists()

        if existing_samples:
            return JsonResponse({
                'error': 'Samples already exist for the selected areas',
                'existing': True
            }, status=400)

        # Create new Sample objects
        # if sample_type == 'Borderline':
        #     return JsonResponse({
        #     'message': 'Samples test borderline approved successfully',
        #     # 'count': len(new_samples)
        # }, status=200)
            
        new_samples = []
        for csv_data in csv_data_objects:
            new_sample = Sample1(
                is_urban=is_urban,
                ben_id=csv_data.id,  # Assuming ben_id is the same as CSVData _id
                sample_type=sample_type,
                key = csv_data.key,
                cp_id=csv_data,
                remarks=remarks,
                created_by=request.user  # Assuming you're using authentication
            )
            new_samples.append(new_sample)

        # Bulk create the new samples
        Sample1.objects.bulk_create(new_samples)

        return JsonResponse({
            'message': 'Samples approved successfully',
            'count': len(new_samples)
        }, status=200)
        
  

class SampledLocations(View):
    def get(self, request):
        try:
            # Base queryset for Sample1 with aggregations
            counts = Sample1.objects.filter(
                sample_type='Regular'
            ).values(
                'cp_id__SB_ao',
                'cp_id__SB_province',
                'cp_id__SB_district',
                'cp_id__SB_area',
                'cp_id__SB_nahia'
            ).annotate(
                sample_count=Count('id'),
                cp_exclusion_error=Count('id', filter=Q(cp_id__vul='No', tpm_records__vul='Yes')),
                cp_inclusion_error=Count('id', filter=Q(cp_id__vul='Yes', tpm_records__vul='No')),
                cp_vul_tpm_vul_null=Count('id', filter=Q(cp_id__vul=None, tpm_records__vul=None)),
                cp_excluded=Count('id', filter=Q(cp_id__exclusion_1=True)),
                tpm_excluded=Count('id', filter=Q(tpm_records__exclusion_1=True)),
                total_tpm=Count('tpm_records__id'),
                tpm_hh_not_found=Count('tpm_records__id', filter=Q(tpm_records__HHFound=False))
            )

            # Perform aggregation on TPM_EE_Data grouped by common fields and 'vul'
            tpmee_counts = TPM_EE_Data.objects.values(
                'SB_ao',
                'SB_province',
                'SB_district',
                'SB_area',
                'SB_nahia',
                'vul'
            ).annotate(
                count=Count('id')
            )

            # Build a mapping from common fields to 'vul' counts
            tpmee_counts_dict = {}
            for item in tpmee_counts:
                key = (
                    item['SB_ao'],
                    item['SB_province'],
                    item['SB_district'],
                    item['SB_area'],
                    item['SB_nahia']
                )
                if key not in tpmee_counts_dict:
                    tpmee_counts_dict[key] = {}
                tpmee_counts_dict[key][item['vul']] = item['count']

            # Convert counts queryset to list
            counts_list = list(counts)

            # Add 'ee' key with TPM_EE_Data counts nested under it
            for item in counts_list:
                key = (
                    item['cp_id__SB_ao'],
                    item['cp_id__SB_province'],
                    item['cp_id__SB_district'],
                    item['cp_id__SB_area'],
                    item['cp_id__SB_nahia']
                )
                ee_counts = tpmee_counts_dict.get(key, {})
                item['ee'] = ee_counts

            # Prepare the result
            result = {
                'counts': counts_list
            }

            # Return JSON response
            return JsonResponse(result, safe=True)

        except Exception as e:
            logger.exception("An error occurred while fetching sampled locations.")
            return JsonResponse(
                {'error': 'An error occurred while processing your request.'},
                status=500
            )
            
from django.db.models import ExpressionWrapper, FloatField, F

class SampledLocations_bordeline(View):
    def get(self, request):
        try:
            # Base queryset for Sample1 with aggregations
            counts = Sample1.objects.filter(
                sample_type='Regular'
            ).values(
                'cp_id__SB_ao',
                'cp_id__SB_province',
                'cp_id__SB_district',
                'cp_id__SB_area',
                'cp_id__SB_nahia'
            ).annotate(
                sample_count=Count('id'),
                cp_inclusion_error=Count('id', filter=Q(cp_id__vul='Yes', tpm_records__vul='No')),
                total_tpm=Count('tpm_records__id'),
                tpm_hh_not_found=Count('tpm_records__id', filter=Q(tpm_records__HHFound=False)),
                inclusion_error_percent=ExpressionWrapper(
                    F('cp_inclusion_error') * 100.0 / F('total_tpm'),
                    output_field=FloatField()
                )
            ).filter(
                inclusion_error_percent__gt=10
            )

           
            # Get list of cp_ids in Sample1
            sampled_cp_ids = Sample1.objects.values_list('cp_id', flat=True)

            # Get counts of CPs matching the criteria and not in sample
            cp_counts = CPDataModel1.objects.filter(
                CP_Calculation__in=[6, 7],
                vul='Yes'
            ).exclude(
                id__in=sampled_cp_ids
            ).values(
                'SB_ao',
                'SB_province',
                'SB_district',
                'SB_area',
                'SB_nahia'
            ).annotate(
                cp_not_in_sample_count=Count('id')
            )

            # Build a mapping from area keys to cp_not_in_sample_counts
            cp_counts_dict = {}
            for item in cp_counts:
                key = (
                    item['SB_ao'] or '',
                    item['SB_province'] or '',
                    item['SB_district'] or '',
                    item['SB_area'] or '',
                    item['SB_nahia'] or ''
                )
                cp_counts_dict[key] = item['cp_not_in_sample_count']

            # Convert counts queryset to list
            counts_list = list(counts)

            # Add 'ee' key with TPM_EE_Data counts nested under it, cp_not_in_sample_count, and estimated_sample
            for item in counts_list:
                key = (
                    item['cp_id__SB_ao'] or '',
                    item['cp_id__SB_province'] or '',
                    item['cp_id__SB_district'] or '',
                    item['cp_id__SB_area'] or '',
                    item['cp_id__SB_nahia'] or ''
                )
              

                cp_not_in_sample_count = cp_counts_dict.get(key, 0)
                item['cp_not_in_sample_count'] = cp_not_in_sample_count

                # Calculate estimated_sample based on the condition
                if cp_not_in_sample_count <= 500:
                    estimated_sample = cp_not_in_sample_count
                else:
                    estimated_sample = 500
                item['estimated_sample'] = estimated_sample

            # Prepare the result
            result = {
                'counts': counts_list
            }

            # Return JSON response
            return JsonResponse(result, safe=True)

        except Exception as e:
            logger.exception("An error occurred while fetching sampled locations.")
            return JsonResponse(
                {'error': 'An error occurred while processing your request.'},
                status=500
            )




import random
from django.db.models import Q

import json
from django.views import View
from django.http import JsonResponse
from django.db.models import Q

class GenerateRandomSites(View):
    def post(self, request):
        try:
            # Parse the JSON data from the request body
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'error': 'Invalid JSON data.'}, status=400)

            # Extract parameters from POST data
            area_office = data.get('area_office')
            province = data.get('province')
            district = data.get('district')
            nahia = data.get('nahia')
            estimated_sample_size = data.get('estimated_sample_size')
            estimated_sample_size = int(estimated_sample_size)
            # print(estimated_sample_size)
            # Validate input parameters
            if nahia == 'null' or nahia is None:
                nahia = None

            # Validate that all required fields are provided
            if not all([area_office, province, district, estimated_sample_size]):
                return JsonResponse({'error': 'Missing required fields.'}, status=400)

            # Validate that estimated_sample_size is a positive integer
            if not isinstance(estimated_sample_size, int) or estimated_sample_size <= 0:
                return JsonResponse({'error': 'Invalid estimated_sample_size provided.'}, status=400)

            # Step 1: Filter cp_data having CP_Calculation 6 and 7
            cp_data = CPDataModel1.objects.filter(
                Q(CP_Calculation=6) | Q(CP_Calculation=7),
                SB_ao=area_office,
                SB_province=province,
                SB_district=district
            ).exclude(
                id__in=Sample1.objects.values_list('cp_id', flat=True)
            )

            # If Nahia is provided, filter by it
            if nahia:
                cp_data = cp_data.filter(SB_nahia=nahia)

            # Count the records with CP_Calculation 6 and 7
            count_6 = cp_data.filter(CP_Calculation=6).count()
            count_7 = cp_data.filter(CP_Calculation=7).count()
            total_count = count_6 + count_7

            selected_sites = []

            if total_count <= 500:
                # If total count is <= 500, include all records not in sample
                selected_sites = list(cp_data)
            else:
                # Randomly select 'estimated_sample_size' records
                selected_sites = list(cp_data.order_by('?')[:estimated_sample_size])
            
            samples = self.prepare_sample_data(selected_sites)
            # Prepare the result
            result = {
                'selected_sites': samples
            }

            # Return JSON response
            return JsonResponse(result, safe=True)

        except Exception as e:
            logger.exception("An error occurred while generating random sites.")
            return JsonResponse(
                {'error': 'An error occurred while processing your request.'},
                status=500
            )
    def prepare_sample_data(self, sample):
        sample_data = []
        for record in sample:
            record_data = {}
            for field in record._meta.fields:
                field_name = field.name
                field_value = getattr(record, field_name)
                # Convert non-JSON serializable types to string
                if isinstance(field_value, (datetime.date, datetime.datetime)):
                    field_value = field_value.isoformat()
                record_data[field_name] = field_value
            sample_data.append(record_data)
        return sample_data



# from django.views import View
# from django.http import JsonResponse
# from django.db.models import Q, OuterRef, Exists, Subquery
# from .models import CSVData, Sample, TPMCSVData
# import json

# class FinalListDataAnalysis(View):
#     def post(self, request):
#         json_data = json.loads(request.body)
#         area_office = json_data.get('area_office')
#         province = json_data.get('province')
#         district = json_data.get('district')
#         nahia = json_data.get('nahia')
#         error_type = json_data.get('error_type')
        
#         if not all([area_office, province, district]):
#             return JsonResponse({'error': 'Missing required parameters'}, status=400)

#         # Base queryset for CSVData
#         csvdata_queryset = CSVData.objects.filter(
#             SB_ao=area_office,
#             SB_province=province,
#             SB_district=district
#         )

#         if nahia:
#             csvdata_queryset = csvdata_queryset.filter(SB_nahia=nahia)

#         # Subquery for Sample
#         sample_subquery = Sample.objects.filter(cp_id=OuterRef('pk'))

#         # Annotate CSVData with Sample information
#         csvdata_queryset = csvdata_queryset.annotate(
#             has_sample=Exists(sample_subquery),
#             sample_type=Subquery(sample_subquery.values('sample_type')[:1])
#         )

#         # Separate common and non-common records
#         common_records = []
#         non_common_records = []

#         for csvdata in csvdata_queryset:
#             record = {field.name: getattr(csvdata, field.name) for field in csvdata._meta.fields}
#             record['has_sample'] = csvdata.has_sample
#             record['sample_type'] = csvdata.sample_type

#             if csvdata.has_sample:
#                 # For records in both CSVData and Sample
#                 tpm_data = TPMCSVData.objects.filter(sample__cp_id=csvdata.id).first()
#                 if tpm_data:
#                     record['tpm_vul'] = tpm_data.vul
#                     record['tpm_HHFound'] = tpm_data.HHFound

#                 if record.get('assessmentType') == 'Replacement Assessment':
#                     if record.get('cp_vul'):
#                         record['remark'] = 'Selected: During replace assessment by CP'
#                     else:
#                         record['remark'] = 'Rejected: During replace assessment by CP'
#                 elif record.get('vul') and record.get('tpm_vul'):
#                     record['remark'] = 'Selected: During spot-check'
#                 elif not record.get('vul') and record.get('tpm_vul'):
#                     record['remark'] = 'Selected: During spot-check, initially not eligible by CP'
#                 elif record.get('vul') and not record.get('tpm_vul'):
#                     record['remark'] = 'Selected: Rejected During Spot-check'
#                 elif not record.get('tpm_HHFound'):
#                     record['remark'] = 'Selected: Rejected During Spot-check due to HH not found'
#                 else:
#                     record['remark'] = 'No specific remark (has sample)'

#                 common_records.append(record)
#             else:
#                 # For records only in CSVData
#                 if record.get('vul'):
#                     record['remark'] = 'Selected: Beneficiary Verification by CP'
#                 else:
#                     record['remark'] = 'Rejected: Beneficiary Verification by CP'

#                 non_common_records.append(record)

#         # Apply error_type filter if provided
#         if error_type and error_type != 'all':
#             error_type_mapping = {
#                 'inclusion': 'Selected',
#                 'exclusion': 'Rejected',
#                 'other': 'No specific remark'
#             }
#             filter_term = error_type_mapping.get(error_type.lower(), error_type)
#             common_records = [r for r in common_records if filter_term in r.get('remark', '')]
#             non_common_records = [r for r in non_common_records if filter_term in r.get('remark', '')]

#         # Prepare the result
#         result = {
#             'common_records': {
#                 'data': common_records,
#                 'count': len(common_records),
#             },
#             'non_common_records': {
#                 'data': non_common_records,
#                 'count': len(non_common_records),
#             },
#             'total_count': len(common_records) + len(non_common_records),
#             'csvdata_count': csvdata_queryset.count(),
#             'query_params': json_data,
#             'debug_info': {
#                 'csvdata_initial_count': CSVData.objects.filter(
#                     SB_ao=area_office,
#                     SB_province=province,
#                     SB_district=district
#                 ).count(),
#                 'csvdata_with_nahia_count': csvdata_queryset.count(),
#                 'sample_count': Sample.objects.filter(cp_id__in=csvdata_queryset).count(),
#                 'tpm_count': TPMCSVData.objects.filter(sample__cp_id__in=csvdata_queryset).count(),
#             }
#         }

#         return JsonResponse(result, safe=False)

import json
import logging
from django.http import JsonResponse, HttpResponseBadRequest
from django.views import View
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from .models import CSVData, TPMCSVData, Sample


class FinalListDataAnalysis(View):

    def post(self, request, *args, **kwargs):
        try:
            json_data = json.loads(request.body)
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON data: {e}")
            return HttpResponseBadRequest("Invalid JSON data")

        area_office = json_data.get('area_office')
        province = json_data.get('province')
        district = json_data.get('district')
        nahia = json_data.get('nahia')
        error_type = json_data.get('error_type')
        print([area_office, province, district, nahia])

        if not all([area_office, province, district, nahia]):
            logger.error("Missing required parameters")
            return HttpResponseBadRequest("Missing required parameters")

        try:
            cp_data = self.get_cp_data(area_office, province, district, nahia)
            sample_data = Sample1.objects.filter(cp_id__in=cp_data).select_related('cp_id')
            tpm_data = TPM_SC_Data.objects.filter(sample__in=sample_data).select_related('sample__cp_id')

            common_data = self.process_common_data(cp_data, tpm_data)
            non_common_data = self.process_non_common_data(cp_data, tpm_data)

            counts = self.calculate_counts(cp_data, tpm_data, common_data, non_common_data)

            return JsonResponse({
                'common_data': common_data,
                'non_common_data': non_common_data,
                'counts': counts
            })

        except Exception as e:
            logger.error(f"Error processing data: {e}")
            return JsonResponse({'error': 'Internal Server Error'}, status=500)

    def get_cp_data(self, area_office, province, district, nahia):
        filters = {
            'SB_ao': area_office,
            'SB_province': province,
            'SB_district': district,
        }
        
        # Only add SB_nahia to the filter if nahia is not 'null' or None
        if nahia is not None and nahia != 'null':
            filters['SB_nahia'] = nahia
        
        return CPDataModel1.objects.filter(**filters)


    def process_common_data(self, cp_data, tpm_data):
        common_data = []
        tpm_data_map = {tpm.sample.cp_id.id: tpm for tpm in tpm_data}

        for cp in cp_data.filter(id__in=tpm_data_map.keys()):
            tpm = tpm_data_map[cp.id]
            result = self.get_all_cp_data(cp)

            result.update({
                'moda_id': cp.moda_id,
                'tpm_vul': tpm.vul,
                'HHFound': tpm.HHFound,
                'TPM_Calculation': tpm.TPM_Calculation,
                'status': self.determine_status(cp, tpm)
            })

            common_data.append(result)
        return common_data

    def process_non_common_data(self, cp_data, tpm_data):
        non_common_data = []
        tpm_cp_ids = {tpm.sample.cp_id.id for tpm in tpm_data}

        for cp in cp_data.exclude(id__in=tpm_cp_ids):
            result = self.get_all_cp_data(cp)
            # result['status'] = self.determine_non_common_status(cp)
            result.update({
                'moda_id': cp.moda_id,
                'tpm_vul': '',
                'HHFound': '',
                'TPM_Calculation':'',
                'status': self.determine_non_common_status(cp)
            })
            non_common_data.append(result)
        return non_common_data

    def determine_status(self, cp, tpm):
        if tpm.HHFound == False:
            return 'Rejected: Due to hh not found during spotcheck'
        if cp.vul == 'Yes' and tpm.vul == 'Yes':
            return 'Selected: Due to vulnerable by CP and TPM during spotcheck'
        if cp.vul == 'Yes' and tpm.vul == 'No':
            return 'Rejected: during spotcheck, initailly selected during cp verification'
        if cp.vul =='No' and tpm.vul == 'Yes':
            return 'Selected: during spotcheck, initailly rejected during cp verification'
        if cp.vul == 'Yes' and tpm.vul == 'Excluded':
            return 'Rejected: Excluded by TPM during spotcheck'
            

    def determine_non_common_status(self, cp):
        if cp.assessmentType != 'Replacement Assessment':
            if cp.vul == 'Excluded':
                return 'Rejected: Exclusion by CP'
            return 'Selected: During CP Verification' if cp.vul == 'Yes' else 'Rejected: During CP Verification'
        if cp.assessmentType == 'Replacement Assessment':
            if cp.vul == 'Excluded':
                return 'Rejected: Exclusion by CP'
        return 'Selected: During replacement Assessment' if cp.vul =='Yes' else 'Rejected: During replacement Assessment'

    def get_all_cp_data(self, cp):
        result = {field.name: getattr(cp, field.name) for field in CPDataModel1._meta.fields}
        result['cp_vul'] = result.pop('vul')
        return result

    def calculate_counts(self, cp_data, tpm_data, common_data, non_common_data):
        total_cp_data = cp_data.count()
        vulnerable_by_cp = cp_data.filter(vul='Yes').count()
        vulnerable_by_tpm = tpm_data.filter(vul='Yes').count()
        cp_border_line = cp_data.filter(CP_Calculation__in=[6,7]).count()
        cp_excluded = cp_data.filter(vul='Excluded').count()
        tpm_excluded = tpm_data.filter(vul='Excluded').count()

        total_selected = sum(
            1 for entry in common_data + non_common_data if 'Selected' in entry['status']
        )
        total_rejected = sum(
            1 for entry in common_data + non_common_data if 'Rejected' in entry['status']
        )

        total_replacement = cp_data.filter(
            assessmentType='Replacement Assessment', vul='Yes'
        ).count()

        percentage_selected = (total_selected / total_cp_data) * 100 if total_cp_data > 0 else 0
        percentage_rejected = (total_rejected / total_cp_data) * 100 if total_cp_data > 0 else 0

        total_unique_tpm_entries = tpm_data.distinct().count()

        total_non_replacement = cp_data.exclude(
            assessmentType='Replacement Assessment'
        ).count()


        rejection_reasons_distribution = {}
        for entry in common_data + non_common_data:
            if 'Rejected' in entry['status']:
                reason = entry['status'].split(':', 1)[-1].strip()
                rejection_reasons_distribution[reason] = rejection_reasons_distribution.get(reason, 0) + 1

        return {
            'Total CP Data': total_cp_data,
            'Total Non-Replacement Assessments': total_non_replacement,
            'Total Replacement Assessments': total_cp_data - total_non_replacement,
            'Vulnerable By CP': vulnerable_by_cp,
            'Vulnerable By TPM': vulnerable_by_tpm,
            'Total Selected': str(total_selected) + f" ({percentage_selected:.1f}%)",
            'Total Rejected': str(total_rejected) + f" ({percentage_rejected:.1f}%)",
            'Total Excluded by CP*': str(cp_excluded) + f" ({(cp_excluded/total_cp_data)*100:.1f}%)",
            'Total Excluded by TPM*': str(tpm_excluded) + f" ({(tpm_excluded/total_unique_tpm_entries)*100:.1f}%)",
            'Selected During Replacement Assessment': total_replacement,
            'Total Spotcheck': total_unique_tpm_entries,
            'Total Borderline (6, 7)': str(cp_border_line) + f" ({(cp_border_line/total_cp_data)*100:.1f}%)",
            '* Based on Exlusion Question':''
            
        }

from django.views import View
from django.http import JsonResponse, HttpResponseBadRequest
from django.forms.models import model_to_dict
import json
import logging

logger = logging.getLogger(__name__)


class FinalListDataAnalysis(View):

    def post(self, request, *args, **kwargs):
        try:
            json_data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError as e:
            logger.error(f"Invalid JSON data: {e}")
            return HttpResponseBadRequest("Invalid JSON data")

        area_office = json_data.get('area_office')
        province = json_data.get('province')
        district = json_data.get('district')
        nahia = json_data.get('nahia')

        logger.debug(f"Received parameters: {area_office}, {province}, {district}, {nahia}")

        if not all([area_office, province, district, nahia]):
            logger.error("Missing required parameters")
            return HttpResponseBadRequest("Missing required parameters")

        try:
            cp_data = self.get_cp_data(area_office, province, district, nahia)
            sample_data = Sample1.objects.filter(cp_id__in=cp_data).select_related('cp_id')
            tpm_data = TPM_SC_Data.objects.filter(sample__in=sample_data).select_related('sample__cp_id')

            common_data = self.process_common_data(cp_data, tpm_data)
            non_common_data = self.process_non_common_data(cp_data, tpm_data)

            counts = self.calculate_counts(cp_data, tpm_data, common_data, non_common_data)

            return JsonResponse({
                'common_data': common_data,
                'non_common_data': non_common_data,
                'counts': counts
            })

        except Exception as e:
            logger.exception("Error processing data")
            return JsonResponse({'error': 'Internal Server Error'}, status=500)

    def get_cp_data(self, area_office, province, district, nahia):
        filters = {
            'SB_ao': area_office,
            'SB_province': province,
            'SB_district': district,
        }

        # Handle 'null' or None values for nahia
        if nahia and nahia.lower() != 'null':
            filters['SB_nahia'] = nahia

        return CPDataModel1.objects.filter(**filters)

    def process_common_data(self, cp_data, tpm_data):
        common_data = []
        # Create a mapping of cp_id to tpm_data for quick lookup
        tpm_data_map = {
            tpm.sample.cp_id.id: tpm for tpm in tpm_data
            if tpm.sample and tpm.sample.cp_id
        }

        cp_with_tpm = cp_data.filter(id__in=tpm_data_map.keys())

        for cp in cp_with_tpm:
            tpm = tpm_data_map.get(cp.id)
            if tpm:
                result = self.get_all_cp_data(cp)
                result.update({
                    'moda_id':cp.moda_id,
                    'tpm_vul': tpm.vul,
                    'HHFound': tpm.HHFound,
                    'TPM_Calculation': tpm.TPM_Calculation,
                    'status': self.determine_status(cp, tpm)
                })
                common_data.append(result)
        return common_data

    def process_non_common_data(self, cp_data, tpm_data):
        non_common_data = []
        tpm_cp_ids = {
            tpm.sample.cp_id.id for tpm in tpm_data
            if tpm.sample and tpm.sample.cp_id
        }

        cp_without_tpm = cp_data.exclude(id__in=tpm_cp_ids)

        for cp in cp_without_tpm:
            result = self.get_all_cp_data(cp)
            result.update({
                'moda_id':cp.moda_id,
                'tpm_vul': '',
                'HHFound': '',
                'TPM_Calculation': '',
                'status': self.determine_non_common_status(cp)
            })
            non_common_data.append(result)
        return non_common_data

    def determine_status(self, cp, tpm):
        if not tpm.HHFound:
            return 'Rejected: Due to HH not found during spotcheck'

        if cp.vul == 'Yes':
            if tpm.vul == 'Yes':
                return 'Selected: Vulnerable by both CP and TPM during spotcheck'
            elif tpm.vul == 'No':
                return 'Rejected: During spotcheck, initially selected during CP verification'
            elif tpm.vul == 'Excluded':
                return 'Rejected: Excluded by TPM during spotcheck based on exclusion question'

        if cp.vul == 'No' and tpm.vul == 'Yes':
            return 'Selected: During spotcheck, initially rejected during CP verification'

        return 'Status Unknown'

    def determine_non_common_status(self, cp):
        if cp.vul == 'Excluded':
            return 'Rejected: Exclusion by CP based on exlusion question'

        if cp.assessmentType == 'Replacement Assessment':
            return ('Selected: During Replacement Assessment' if cp.vul == 'Yes'
                    else 'Rejected: During Replacement Assessment')

        return ('Selected: During CP Verification' if cp.vul == 'Yes'
                else 'Rejected: During CP Verification')

    def get_all_cp_data(self, cp):
        cp_dict = model_to_dict(cp)
        cp_dict['cp_vul'] = cp_dict.pop('vul', '')
        return cp_dict

    def calculate_counts(self, cp_data, tpm_data, common_data, non_common_data):
        total_cp_data = cp_data.count()
        vulnerable_by_cp = cp_data.filter(vul='Yes').count()
        vulnerable_by_tpm = tpm_data.filter(vul='Yes').count()
        cp_borderline = cp_data.filter(CP_Calculation__in=[6, 7]).count()

        total_selected = sum(
            1 for entry in common_data + non_common_data if 'Selected' in entry.get('status', '')
        )
        total_rejected = sum(
            1 for entry in common_data + non_common_data if 'Rejected' in entry.get('status', '')
        )

        total_replacement = cp_data.filter(
            assessmentType='Replacement Assessment', vul='Yes'
        ).count()

        percentage_selected = (total_selected / total_cp_data) * 100 if total_cp_data > 0 else 0
        percentage_rejected = (total_rejected / total_cp_data) * 100 if total_cp_data > 0 else 0

        total_unique_tpm_entries = tpm_data.values('sample__cp_id').distinct().count()

        total_non_replacement = cp_data.exclude(
            assessmentType='Replacement Assessment'
        ).count()

        # Calculate rejection reasons distribution
        rejection_reasons_distribution = {}
        for entry in common_data + non_common_data:
            status = entry.get('status', '')
            if 'Rejected' in status:
                reason = status.split(':', 1)[-1].strip()
                rejection_reasons_distribution[reason] = rejection_reasons_distribution.get(reason, 0) + 1

        counts = {
            'Total CP Data': total_cp_data,
            'Total Non-Replacement Assessments': total_non_replacement,
            'Total Replacement Assessments': total_cp_data - total_non_replacement,
            'Vulnerable By CP': vulnerable_by_cp,
            'Vulnerable By TPM': vulnerable_by_tpm,
            'Total Selected': f"{total_selected} ({percentage_selected:.1f}%)",
            'Total Rejected': f"{total_rejected} ({percentage_rejected:.1f}%)",
            'Selected During Replacement Assessment': total_replacement,
            'Total Spotcheck': total_unique_tpm_entries,
            'Total Borderline (6, 7)': f"{cp_borderline} ({(cp_borderline / total_cp_data) * 100 if total_cp_data else 0:.1f}%)",
            'Rejection Reasons Given Bellow':'',
            'Reasons': rejection_reasons_distribution
        }

        return counts


class FinalListApproval(View):
    
    def post(self, request, *args, **kwargs):
        try:
            json_data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return HttpResponseBadRequest("Invalid JSON data")

        data = json_data.get('data')
        final_comment = json_data.get('final_comment')

        if not data or not final_comment:
            return HttpResponseBadRequest("Missing required data")

        try:
            with transaction.atomic():
                for item in data:
                    # print(item)
                    try:
                        # Fetching the CPDataModel1 instance based on id
                        bs = CPDataModel1.objects.get(id=item['id'])
                        # print(bs)
                        # Creating a FinalApproval entry
                        FinalApproval.objects.create(
                            bs=bs,
                            bs_key=bs.key,  # Using dot notation to access model field 'key'
                            final_comment=final_comment,
                            tpm_vul=item.get('tpm_vul', ''),  # Using get to avoid KeyError if field doesn't exist
                            tpm_hh_found=item.get('HHFound', False),  # Assuming HHFound is a boolean field
                            status=item.get('status', ''),  # Defaulting to empty string if status is not provided
                            created_by=request.user
                        )
                    except CPDataModel1.DoesNotExist:
                        print(f"CPDataModel with id {item['id']} not found.")
                        return HttpResponseBadRequest(f"CPDataModel with id {item['id']} not found.")
        except Exception as e:
            print(f"An unexpected error occurred: {str(e)}")
            return HttpResponseBadRequest(f"An unexpected error occurred: {str(e)}")

        return JsonResponse({'status': 'success'}, status=200)

        




import openpyxl
from django.http import StreamingHttpResponse, HttpResponseBadRequest
from django.apps import apps
from io import BytesIO
import uuid
import datetime

def export_model_to_excel(request, model_name):
    try:
        # Get the model based on the provided model name
        model = apps.get_model('base', model_name)
        if model is None:
            return HttpResponseBadRequest("Invalid model name provided.")

        # Create an Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = model_name

        # Get the model's fields
        fields = [field.name for field in model._meta.fields]

        # Write the header row
        for col_num, field in enumerate(fields, 1):
            worksheet.cell(row=1, column=col_num, value=field)

        # Write the data rows
        for row_num, obj in enumerate(model.objects.all(), 2):
            for col_num, field in enumerate(fields, 1):
                value = getattr(obj, field)
                
                # Convert non-convertible types to string
                if isinstance(value, (uuid.UUID, datetime.date, datetime.datetime)):
                    value = str(value)
                elif not isinstance(value, (int, float, str)):
                    value = str(value)  # Convert any other complex object to string
                
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Save the workbook to a bytes buffer
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Create a streaming response
        response = StreamingHttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={model_name}.xlsx'
        return response

    except LookupError:
        return HttpResponseBadRequest("Model not found. Please provide a valid model name.")

def check_keys_in_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            return HttpResponse("No file uploaded.", status=400)

        try:
            # Read the Excel file into a pandas DataFrame
            df = pd.read_excel(excel_file)

            # Ensure 'KEY' column exists in the DataFrame
            if 'KEY' not in df.columns:
                return HttpResponse("Excel file must contain a 'KEY' column.", status=400)

            # Fetch all keys from the model
            model_keys = set(CPDataModel1.objects.values_list('key', flat=True))

            # Add a 'Match' column based on whether 'KEY' exists in model_keys
            df['Match'] = df['KEY'].apply(lambda x: x in model_keys)

            # Save the modified DataFrame to an Excel file in memory
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            output.seek(0)

            # Prepare the response with the modified Excel file
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=checked_keys.xlsx'

            return response

        except Exception as e:
            return HttpResponse(f"An error occurred: {str(e)}", status=500)

    else:
        # Render the upload form template for GET requests
        return render(request, 'check_upload_excel.html')

# Set up logging
logger = logging.getLogger(__name__)

def get_grouped_data(queryset, group_fields, count_field, prefix=''):
    """
    Utility function to annotate queryset with counts and return values as dictionaries.
    
    Args:
        queryset: Django queryset to group and annotate.
        group_fields: List of fields to group by.
        count_field: Field to count on the queryset.
        prefix: (Optional) prefix for field names when using related fields.
    
    Returns:
        list of grouped dictionaries.
    """
    try:
        data = (
            queryset
            .values(*group_fields)
            .annotate(**{count_field: Count('id')})
        )
        return list(data)
    except Exception as e:
        logger.error(f"Error fetching grouped data: {e}")
        return []

def combine_grouped_data(cp_data, sample_data, tpm_ee_data, tpm_sc_data):
    """
    Combines data from different models into one structure based on shared keys.
    
    Args:
        cp_data: Data from CPDataModel1 grouped by common fields.
        sample_data: Data from Sample1 grouped by common fields.
        tpm_ee_data: Data from TPM_EE_Data grouped by common fields.
        tpm_sc_data: Data from TPM_SC_Data grouped by common fields.
    
    Returns:
        A list of dictionaries with combined counts.
    """
    combined_data = {}

    # Combine CP data
    for cp in cp_data:
        key = (
            cp['SB_ao'], cp['SB_B_1'], cp['SB_province'], 
            cp['SB_B_2'], cp['SB_district'], cp['SB_area'], cp['SB_nahia']
        )
        combined_data[key] = {
            'SB_ao': cp['SB_ao'],
            'SB_B_1': cp['SB_B_1'],
            'SB_province': cp['SB_province'],
            'SB_B_2': cp['SB_B_2'],
            'SB_district': cp['SB_district'],
            'SB_area': cp['SB_area'],
            'SB_nahia': cp['SB_nahia'],
            'cp_count': cp['cp_count'],
            'sample_count': 0,
            'tpm_ee_count': 0,
            'tpm_sc_count': 0
        }

    # Combine Sample data
    for sample in sample_data:
        key = (
            sample['cp_id__SB_ao'], sample['cp_id__SB_B_1'], 
            sample['cp_id__SB_province'], sample['cp_id__SB_B_2'], 
            sample['cp_id__SB_district'], sample['cp_id__SB_area'], sample['cp_id__SB_nahia']
        )
        if key in combined_data:
            combined_data[key]['sample_count'] = sample['sample_count']
        else:
            combined_data[key] = {
                'SB_ao': sample['cp_id__SB_ao'],
                'SB_B_1': sample['cp_id__SB_B_1'],
                'SB_province': sample['cp_id__SB_province'],
                'SB_B_2': sample['cp_id__SB_B_2'],
                'SB_district': sample['cp_id__SB_district'],
                'SB_area': sample['cp_id__SB_area'],
                'SB_nahia': sample['cp_id__SB_nahia'],
                'cp_count': 0,
                'sample_count': sample['sample_count'],
                'tpm_ee_count': 0,
                'tpm_sc_count': 0
            }

    # Combine TPM EE data
    for tpm_ee in tpm_ee_data:
        key = (
            tpm_ee['SB_ao'], tpm_ee['SB_B_1'], tpm_ee['SB_province'], 
            tpm_ee['SB_B_2'], tpm_ee['SB_district'], tpm_ee['SB_area'], tpm_ee['SB_nahia']
        )
        if key in combined_data:
            combined_data[key]['tpm_ee_count'] = tpm_ee['tpm_ee_count']
        else:
            combined_data[key] = {
                'SB_ao': tpm_ee['SB_ao'],
                'SB_B_1': tpm_ee['SB_B_1'],
                'SB_province': tpm_ee['SB_province'],
                'SB_B_2': tpm_ee['SB_B_2'],
                'SB_district': tpm_ee['SB_district'],
                'SB_area': tpm_ee['SB_area'],
                'SB_nahia': tpm_ee['SB_nahia'],
                'cp_count': 0,
                'sample_count': 0,
                'tpm_ee_count': tpm_ee['tpm_ee_count'],
                'tpm_sc_count': 0
            }

    # Combine TPM SC data
    for tpm_sc in tpm_sc_data:
        key = (
            tpm_sc['sample__cp_id__SB_ao'], tpm_sc['sample__cp_id__SB_B_1'], 
            tpm_sc['sample__cp_id__SB_province'], tpm_sc['sample__cp_id__SB_B_2'], 
            tpm_sc['sample__cp_id__SB_district'], tpm_sc['sample__cp_id__SB_area'], tpm_sc['sample__cp_id__SB_nahia']
        )
        if key in combined_data:
            combined_data[key]['tpm_sc_count'] = tpm_sc['tpm_sc_count']
        else:
            combined_data[key] = {
                'SB_ao': tpm_sc['sample__cp_id__SB_ao'],
                'SB_B_1': tpm_sc['sample__cp_id__SB_B_1'],
                'SB_province': tpm_sc['sample__cp_id__SB_province'],
                'SB_B_2': tpm_sc['sample__cp_id__SB_B_2'],
                'SB_district': tpm_sc['sample__cp_id__SB_district'],
                'SB_area': tpm_sc['sample__cp_id__SB_area'],
                'SB_nahia': tpm_sc['sample__cp_id__SB_nahia'],
                'cp_count': 0,
                'sample_count': 0,
                'tpm_ee_count': 0,
                'tpm_sc_count': tpm_sc['tpm_sc_count']
            }

    # Return as list
    return list(combined_data.values())

def summary_view(request):
    try:
        # Fetch data from models and group by common fields
        group_fields = ['SB_ao', 'SB_B_1', 'SB_province', 'SB_B_2', 'SB_district', 'SB_area', 'SB_nahia']
        
        # Group CPDataModel1
        cp_data = get_grouped_data(CPDataModel1.objects, group_fields, 'cp_count')
        
        # Group Sample1
        sample_data = get_grouped_data(
            Sample1.objects.select_related('cp_id'), 
            [f'cp_id__{field}' for field in group_fields], 
            'sample_count'
        )
        
        # Group TPM_SC_Data
        tpm_sc_data = get_grouped_data(
            TPM_SC_Data.objects.select_related('sample__cp_id'), 
            [f'sample__cp_id__{field}' for field in group_fields], 
            'tpm_sc_count'
        )
        
        # Group TPM_EE_Data
        tpm_ee_data = get_grouped_data(TPM_EE_Data.objects, group_fields, 'tpm_ee_count')

        # Combine the data
        combined_data = combine_grouped_data(cp_data, sample_data, tpm_ee_data, tpm_sc_data)

        # Return JSON response
        return JsonResponse({'counts': combined_data}, status=200)
    
    except Exception as e:
        logger.error(f"Error generating summary: {e}")
        return JsonResponse({'error': 'An error occurred while generating the summary'}, status=500)
    
class summary_view(View):
    def get(self, request):
        try:
            # Base queryset for Sample1 with necessary aggregations
            counts = Sample1.objects.values(
                'cp_id__SB_ao',
                'cp_id__SB_province',
                'cp_id__SB_district',
                'cp_id__SB_area',
                'cp_id__SB_nahia'
            ).annotate(
                sample_count=Count('id'),  # Count of samples
                cp_inclusion_error=Count('id', filter=Q(cp_id__vul='Yes', tpm_records__vul='No')),  # Inclusion error count
                total_tpm=Count('tpm_records__id'),  # Total TPM records
                tpm_hh_not_found=Count('tpm_records__id', filter=Q(tpm_records__HHFound=False)),  # Count of HH not found
                inclusion_error_percent=ExpressionWrapper(
                    F('cp_inclusion_error') * 100.0 / F('total_tpm'),
                    output_field=FloatField()
                ),  # Inclusion error as percentage
            )

            # Aggregation on TPM_EE_Data for exclusion error
            tpmee_counts = TPM_EE_Data.objects.values(
                'SB_ao',
                'SB_province',
                'SB_district',
                'SB_area',
                'SB_nahia'
            ).annotate(
                ee_vul_yes_count=Count('id', filter=Q(vul='Yes')),  # Count where ee_vul is 'Yes'
                total_ee_count=Count('id')  # Total EE count
            )

            # Create a dictionary to map EE counts for calculating exclusion error
            tpmee_counts_dict = {}
            for item in tpmee_counts:
                key = (
                    item['SB_ao'],
                    item['SB_province'],
                    item['SB_district'],
                    item['SB_area'],
                    item['SB_nahia']
                )
                tpmee_counts_dict[key] = {
                    'ee_vul_yes_count': item['ee_vul_yes_count'],
                    'total_ee_count': item['total_ee_count']
                }

            # Aggregation on CPDataModel1 to get cp_count
            cp_counts = CPDataModel1.objects.values(
                'SB_ao',
                'SB_province',
                'SB_district',
                'SB_area',
                'SB_nahia'
            ).annotate(
                cp_count=Count('id'),
                cp_selected = Count('id', filter=Q(vul='Yes'))   # Count of selected CPs
            )

            # Create a dictionary to map CP counts
            cp_counts_dict = {}
            for cp_item in cp_counts:
                cp_key = (
                    cp_item['SB_ao'],
                    cp_item['SB_province'],
                    cp_item['SB_district'],
                    cp_item['SB_area'],
                    cp_item['SB_nahia']
                )
                cp_counts_dict[cp_key] = {
                    'cp_count': cp_item['cp_count'],
                    'cp_selected': cp_item['cp_selected']
                }
                

            # Convert counts queryset to list for modification
            counts_list = list(counts)

            # Calculate and append exclusion error based on TPM_EE_Data counts and cp_count from CPDataModel1
            for item in counts_list:
                key = (
                    item['cp_id__SB_ao'],
                    item['cp_id__SB_province'],
                    item['cp_id__SB_district'],
                    item['cp_id__SB_area'],
                    item['cp_id__SB_nahia']
                )
                # Get EE data for exclusion error calculation
                ee_data = tpmee_counts_dict.get(key, {})
                if ee_data.get('total_ee_count', 0) > 0:
                    exclusion_error = ee_data['ee_vul_yes_count'] / ee_data['total_ee_count']
                else:
                    exclusion_error = 0  # Default to 0 if no EE data

                item['exclusion_error'] = exclusion_error * 100
                item['total_ee_count'] = ee_data.get('total_ee_count', 0)

                # Get cp_count from CPDataModel1
                item['cp_count'] = cp_counts_dict.get(key, 0)

            # Prepare the result
            result = {
                'counts': counts_list
            }

            # Return JSON response
            return JsonResponse(result, safe=True)

        except Exception as e:
            logger.exception("An error occurred while fetching sampled locations.")
            return JsonResponse(
                {'error': 'An error occurred while processing your request.'},
                status=500
            )





