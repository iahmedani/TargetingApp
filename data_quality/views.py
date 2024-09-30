import pandas as pd
from django.http import HttpResponse, StreamingHttpResponse
from django.shortcuts import redirect, render
from django.views import View
from django.core.files.storage import default_storage, FileSystemStorage
from django.core.files.base import ContentFile
from dtale.views import startup
import io
import dtale.global_state as global_state
from django.conf import settings
from django.contrib import messages
import os
from datetime import datetime
from openpyxl import load_workbook


from .data_quality_utils import (
    read_dataset,
    find_duplicate_id_number,
    find_duplicate_mobile_number,
    find_q5_a5_error,
    find_q6_a6_error,
    find_child_under_5_error,
    find_potential_hoh_duplicates,
    add_error_remarks,
    merge_error_subsets,
    produce_final_excel
)
from io import BytesIO
import xlsxwriter  # Ensure xlsxwriter is installed

global_state.set_app_settings(dict(enable_custom_filters=True, optimize_dataframe=True, ignore_duplicate=True,  lock_header_menu=True, main_title = 'Tageting DQA App'))


def index(request):
    return HttpResponse("""
        <h1>Django App with D-Tale</h1>
        <p><a href="/data_quality/create-df">Generate sample dataframe in D-Tale</a></p>
        <p><a href="/data_quality/import-data">Import CSV file to D-Tale</a></p>
    """)
    
def create_df(request):
    df = pd.DataFrame(dict(a=[1, 2, 3], b=[4, 5, 6]))
    instance = startup(data=df, ignore_duplicate=True)
    return redirect(f"/flask/dtale/main/{instance._data_id}")

# create data upload csv file and display in D-Tale

class ImportDataView(View):
    def get(self, request):
        return render(request, 'upload.html')
    
    def post(self, request):
        if 'file' not in request.FILES:
            return HttpResponse("No file uploaded", status=400)
        
        file = request.FILES['file']
        
        # Save the file temporarily
        path = default_storage.save('tmp/'+file.name, ContentFile(file.read()))
        
        try:
            # Read the CSV file
            with default_storage.open(path) as f:
                df = pd.read_csv(io.StringIO(f.read().decode('utf-8')))
            
            # Start D-Tale with the imported data
            instance = startup(data=df)
            # print(instance.)
            url = f"/flask/dtale/main/{instance._data_id}"
            print(url)
            # Redirect to D-Tale
            return render(request, 'dtale.html', context={'url': url})
        except Exception as e:
            return HttpResponse(f"Error processing file: {str(e)}", status=400)
        finally:
            # Clean up the temporary file
            default_storage.delete(path)
            
def dqa_analysis(request):
    return render(request, 'dtale.html', context={'url': request.GET.get('url')})

def excel_upload_view(request):
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'No file selected')
            return redirect('upload_excel')
        
        excel_dqa_file = request.FILES['excel_file']
        
        # Check if the uploaded file is an Excel (.xlsx) file
        if not excel_dqa_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload a valid .xlsx file')
            return redirect('upload_excel')

        # Generate a unique filename by appending the current timestamp
        original_filename = excel_dqa_file.name
        file_extension = os.path.splitext(original_filename)[1]  # Get the file extension
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")  # Get the current timestamp
        unique_filename = f"{os.path.splitext(original_filename)[0]}_{timestamp}{file_extension}"

        # Define the folder to save the file in
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'excel_files'))
        
        # Save the file with the unique filename
        filename = fs.save(unique_filename, excel_dqa_file)
        file_url = fs.url(filename)

        # Display success message
        messages.success(request, f'File uploaded successfully: {file_url}')
        
        context = {'file_url': file_url}
        return render(request, 'upload_excel.html', context=context)
        
        
    # Render the file upload form if the request method is GET
    return render(request, 'upload_excel.html')

def std_data_analysis1(request):
    file_url = request.GET.get('file_url')
    
    if file_url:
        # Extract the filename from the file URL
        file_name = os.path.basename(file_url)
        
        # Construct the full file path based on MEDIA_ROOT
        file_path = os.path.join(settings.MEDIA_ROOT, 'excel_files', file_name)
        print(file_path)
        
        if not os.path.exists(file_path):
            return HttpResponse("File not found", status=404)

        try:
            # Use a generator to stream the data in chunks manually
            def excel_data_generator(file_path):
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                
                # Read rows in chunks manually (adjust as needed)
                chunk_size = 1000
                rows = ws.iter_rows(values_only=True)
                chunk = []
                for idx, row in enumerate(rows):
                    chunk.append(row)
                    if (idx + 1) % chunk_size == 0:
                        yield pd.DataFrame(chunk).to_json(orient='records')
                        chunk = []
                
                # Yield the remaining rows if any
                if chunk:
                    yield pd.DataFrame(chunk).to_json(orient='records')

            # Create a StreamingHttpResponse for large data
            response = StreamingHttpResponse(excel_data_generator(file_path), content_type='application/json')
            response['Content-Disposition'] = f'attachment; filename="{file_name}"'

            return response

        except Exception as e:
            return HttpResponse(f"Error processing the file: {str(e)}", status=500)
    else:
        return HttpResponse("No file URL provided", status=400)
    

def std_data_analysis(request):
    file_url = request.GET.get('file_url')
    
    if file_url:
        # Extract the filename from the file URL
        file_name = os.path.basename(file_url)
        
        # Construct the full file path based on MEDIA_ROOT
        file_path = os.path.join(settings.MEDIA_ROOT, 'excel_files', file_name)
        
        df = read_dataset(file_path)

        # Section Two: Duplicate id_number
        dup_id_num = find_duplicate_id_number(df)
        dup_id_num = add_error_remarks(
            dup_id_num,
            'Duplicate ID Number:',
            'Duplicate ID number identified based on paper or electronic tazkira (column name id_number). Please review the entries and determine if the data is valid, or if the duplicate records should be removed.'
        )

        # Section Three: Duplicate Mobile Number
        dup_mob_rows = find_duplicate_mobile_number(df)
        dup_mob_rows = add_error_remarks(
            dup_mob_rows,
            'Duplicate Mobile Number',
            'Duplicate mobile number detected (column name mob). Please review and determine whether the mobile numbers need correction or if any duplicates should be removed.'
        )

        # Section Four: Q5/A5 error
        a5_error = find_q5_a5_error(df)
        a5_error = add_error_remarks(
            a5_error,
            'Vulnerability Q5 Error: (The head of the household has a disability.)',
            'Inconsistency found in Vulnerability Question 5 (column A5). The household is marked as disabled (column HH_head has value 3), but Question 5 indicates "No". Please review the data and correct this discrepancy.'
        )

        # Section Five: Q6/A6 Error
        a6_error = find_q6_a6_error(df)
        a6_error = add_error_remarks(
            a6_error,
            'Vulnerability Q6 Error: (The household has more than 3 children under the age of 5.)',
            'Vulnerability Question 6 has a potential issue: either Q6 (column name A6 is marked 1) is marked "Yes" but there are fewer (in column child_5Num) than 4 children under 5, or Q6 is marked "No" but there are more than 3 children under 5. Please verify the data and make necessary corrections.'
        )

        # Section Six: Children Under 5 Error
        child_5_error = find_child_under_5_error(df)
        child_5_error = add_error_remarks(
            child_5_error,
            'Child Under 5 Error (Do you have  children under 5 in this HH? )',
            'Discrepancy found: The household is marked as having a child under 5 (column name child_5 has value 1), but the number of children under 5 is recorded as 0. Please verify and correct this error in the data.'
        )

        # Section Seven: Potential Head of Household Duplication
        duplicate_rows = find_potential_hoh_duplicates(df)
        if not duplicate_rows.empty:
            duplicate_rows = add_error_remarks(
                duplicate_rows,
                'Potential Head of Household Duplicate',
                'Possible duplicate head of household found based on Beneficiary Name, Father Name, and ID Number. Please review the data to confirm whether it is a valid entry or if duplicates need to be corrected or removed.'
            )

        # Section Eight: Merge Errors
        subsets = [dup_id_num, dup_mob_rows, a5_error, a6_error, child_5_error]
        if not duplicate_rows.empty:
            subsets.append(duplicate_rows)
        merged_errors = merge_error_subsets(subsets)

        # Section Nine: Produce final excel workbook
        df_with_errors, df_without_errors = produce_final_excel(df, merged_errors)

        # Create an in-memory output file for the new workbook
        output = BytesIO()

        # Use a context manager to handle the Excel writer
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            # Write each dataframe to a different worksheet
            df_with_errors.to_excel(writer, sheet_name='Errors', index=False)
            df_without_errors.to_excel(writer, sheet_name='No_Errors', index=False)
            # No need to call writer.close(); it's handled by the context manager

        # Rewind the buffer
        output.seek(0)

        # Set up the HttpResponse
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        processed_file_name = f'{file_name.strip(".xlsx")}_processed.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{processed_file_name}"'

        return response
    else:
        return HttpResponse("No file URL provided", status=400)


def cfac_excel_upload_view(request):
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'No file selected')
            return redirect('upload_excel')
        
        excel_dqa_file = request.FILES['excel_file']
        
        # Check if the uploaded file is an Excel (.xlsx) file
        if not excel_dqa_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload a valid .xlsx file')
            return redirect('upload_excel')

        # Generate a unique filename by appending the current timestamp
        original_filename = excel_dqa_file.name
        file_extension = os.path.splitext(original_filename)[1]  # Get the file extension
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")  # Get the current timestamp
        unique_filename = f"{os.path.splitext(original_filename)[0]}_{timestamp}{file_extension}"

        # Define the folder to save the file in
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'excel_files'))
        
        # Save the file with the unique filename
        filename = fs.save(unique_filename, excel_dqa_file)
        file_url = fs.url(filename)

        # Display success message
        messages.success(request, f'File uploaded successfully: {file_url}')
        
        context = {'file_url': file_url}
        return render(request, 'upload_excel_cfac.html', context=context)
        
        
    # Render the file upload form if the request method is GET
    return render(request, 'upload_excel_cfac.html')



from .cfac_village_list_utils import (
    load_excel_file,
    clean_dataframe,
    check_required_columns,
    ensure_remarks_column,
    map_province_and_district_codes,
    add_remarks_for_missing_codes,
    check_missing_focal_point_contacts,
    create_cfac_codes,
    create_cfac_list,
    create_village_codes,
    create_village_list,
    save_to_excel,
)

def process_excel_file(request):
    file_url = request.GET.get('file_url')
    if file_url:
        # Extract the filename from the file URL
        file_name = os.path.basename(file_url)

        # Construct the full file path based on MEDIA_ROOT
        file_path = os.path.join(settings.MEDIA_ROOT, 'excel_files', file_name)

        # Step 1: Load the Excel file
        df = load_excel_file(file_path)

        # Step 2: Clean the DataFrame
        df = clean_dataframe(df)

        # Step 3: Check for required columns
        required_columns = ['AO', 'Area', 'Province', 'District', 'Village', 'Nahia', 'CFAC Name',
                            'CFAC FP 1', 'FP1 Number', 'CFAC FP2', 'FP2 Number', 'Remarks']
        columns_present, missing_columns = check_required_columns(df, required_columns)
        if not columns_present:
            return HttpResponse(f"Missing required columns: {', '.join(missing_columns)}", status=400)

        # Ensure 'Remarks' column exists
        df = ensure_remarks_column(df)

        # Step 4: Map province and district codes
        province_file = os.path.join(settings.MEDIA_ROOT, 'data_files', 'province.csv')
        district_file = os.path.join(settings.MEDIA_ROOT, 'data_files', 'district.csv')
        df = map_province_and_district_codes(df, province_file, district_file)

        # Step 5: Add remarks for missing codes
        df = add_remarks_for_missing_codes(df)

        # Step 6: Check for missing focal point contacts
        df = check_missing_focal_point_contacts(df)

        # Step 7: Separate valid and error rows
        df['Remarks'] = df['Remarks'].fillna('').astype(str)
        df_valid = df[df['Remarks'] == '']
        df_errors = df[df['Remarks'] != '']

        # Step 8: Separate data by area type
        urban_df = df_valid[df_valid['Area'] == 'Urban Area'].copy()
        rural_df = df_valid[df_valid['Area'] == 'Rural Area'].copy()

        # Step 9: Create CFAC codes
        urban_df = create_cfac_codes(urban_df, is_urban=True)
        rural_df = create_cfac_codes(rural_df, is_urban=False)

        # Step 10: Create CFAC list
        cfac_df = pd.concat([urban_df, rural_df], ignore_index=True)
        cfac_list = create_cfac_list(cfac_df)

        # Step 11: Create village codes
        urban_df = create_village_codes(urban_df, is_urban=True)
        rural_df = create_village_codes(rural_df, is_urban=False)

        # Step 12: Create village list
        village_df = pd.concat([urban_df, rural_df], ignore_index=True)
        village_list = create_village_list(village_df)

        # Step 13: Prepare data for Excel output
        output = BytesIO()
        output_file_name = f"{os.path.splitext(file_name)[0]}_output.xlsx"
        df_dict = {
            'cfac_list': cfac_list,
            'village_list': village_list,
            'original_data': df,
            'errors': df_errors,
        }

        # Save to Excel
        save_to_excel(df_dict, output)

        # Set up the HttpResponse
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{output_file_name}"'

        return response
    else:
        return HttpResponse("No file URL provided", status=400)


def deduplicate_upload(request):
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'No file selected')
            return redirect('upload_excel')
        
        excel_dqa_file = request.FILES['excel_file']
        
        # Check if the uploaded file is an Excel (.xlsx) file
        if not excel_dqa_file.name.endswith('.xlsx'):
            messages.error(request, 'Please upload a valid .xlsx file')
            return redirect('upload_excel')

        # Generate a unique filename by appending the current timestamp
        original_filename = excel_dqa_file.name
        file_extension = os.path.splitext(original_filename)[1]  # Get the file extension
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")  # Get the current timestamp
        unique_filename = f"{os.path.splitext(original_filename)[0]}_{timestamp}{file_extension}"

        # Define the folder to save the file in
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'excel_files'))
        
        # Save the file with the unique filename
        filename = fs.save(unique_filename, excel_dqa_file)
        file_url = fs.url(filename)

        # Display success message
        messages.success(request, f'File uploaded successfully: {file_url}')
        
        context = {'file_url': file_url}
        return render(request, 'deduplicate.html', context=context)
        
        
    # Render the file upload form if the request method is GET
    return render(request, 'deduplicate.html')


def deduplicate_export(request):
    file_url = request.GET.get('file_url')
    if file_url:
        # Extract the filename from the file URL
        file_name = os.path.basename(file_url)

        # Construct the full file path based on MEDIA_ROOT
        file_path = os.path.join(settings.MEDIA_ROOT, 'excel_files', file_name)

        # Step 1: Load the Excel file
        df = load_excel_file(file_path)
        
        df_deduplicated = df.drop_duplicates()

       
        # Step 13: Prepare data for Excel output
        output = BytesIO()
        output_file_name = f"{os.path.splitext(file_name)[0]}_output.xlsx"
        df_dict = {
            'deduplicate_data': df_deduplicated,
            'original_data': df
        }

        # Save to Excel
        save_to_excel(df_dict, output)

        # Set up the HttpResponse
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{output_file_name}"'

        return response
    else:
        return HttpResponse("No file URL provided", status=400)


